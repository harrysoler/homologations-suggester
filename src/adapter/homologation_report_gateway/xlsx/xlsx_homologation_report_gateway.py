import shutil
from datetime import date
from pathlib import Path

from openpyxl import load_workbook

from entities import ApprovedSubject, HomologationReportData
from gateways import HomologationReportGateway
from shared_types import ReportGenerationResult

REPORTS_DIR = Path("homologation_reports")

DEFAULT_FORMATION_LEVEL = "Pregrado"
DEFAULT_IDENTIFICATION_TYPE = "Cédula de Ciudadania"
DEFAULT_HOMOLOGATION_TYPE = "Transición Plan de Estudio"

INSTITUTION = "Universidad Santo Tomás"
CITY = "Tunja"
COUNTRY = "Colombia"
CAREER = "Ingeniería de Sistemas"

ORIGIN_PENSUM = "2018-2"
TARGET_PENSUM = "2026-2"

SUBJECTS_TABLE_START_ROW = 51

ORIGIN_SUBJECTS_TABLE_CODE_COL = 'D'
ORIGIN_SUBJECTS_TABLE_NAME_COL = 'E'
ORIGIN_SUBJECTS_TABLE_CREDITS_COL = 'F'
ORIGIN_SUBJECTS_TABLE_GRADE_COL = 'G'

TARGET_SUBJECTS_TABLE_CODE_COL = 'J'
TARGET_SUBJECTS_TABLE_NAME_COL = 'K'
TARGET_SUBJECTS_TABLE_CREDITS_COL = 'L'
TARGET_SUBJECTS_TABLE_GRADE_COL = 'M'

TOTAL_CREDITS = 129

class XLSXHomologationReportGateway(HomologationReportGateway):
    # TODO: Change template path type to Path
    _template_path: str

    def __init__(self, template_path: str):
        self._template_path = template_path

    @staticmethod
    def _build_filename(student_name: str) -> ReportGenerationResult:
        safe_name = student_name.lower().replace(' ', '_')
        return f'{safe_name}.xlsx'

    def generate_report(self, student: HomologationReportData) -> str:
        REPORTS_DIR.mkdir(exist_ok=True)
        output_path: Path = REPORTS_DIR / self._build_filename(student.name)

        shutil.copy2(self._template_path, output_path)

        wb = load_workbook(output_path)
        ws = wb.active
        ws['L18'] = date.today()

        ws['E20'] = student.name.title()
        ws['K20'] = DEFAULT_FORMATION_LEVEL
        ws['E22'] = DEFAULT_IDENTIFICATION_TYPE
        ws['K22'] = student.identification

        ws['K26'] = DEFAULT_HOMOLOGATION_TYPE

        ws['E29'] = INSTITUTION
        ws['E33'] = CITY
        ws['E35'] = COUNTRY
        ws['E37'] = CAREER
        ws['E39'] = ORIGIN_PENSUM

        ws['K29'] = INSTITUTION
        ws['K33'] = CITY
        ws['K35'] = COUNTRY
        ws['K37'] = CAREER
        ws['K39'] = TARGET_PENSUM

        ws['M43'] = TOTAL_CREDITS
        ws['M45'] = student.get_total_approved_credits()

        self._fill_subjects(ws, student.approved_subjects)

        wb.save(output_path)

        return str(output_path)

    def _fill_subjects(self, ws, subjects: list[ApprovedSubject]):
        current_row = SUBJECTS_TABLE_START_ROW

        for approved in subjects:
            ws[f'{ORIGIN_SUBJECTS_TABLE_CODE_COL}{current_row}'] = approved.origin_subject.code
            ws[f'{ORIGIN_SUBJECTS_TABLE_NAME_COL}{current_row}'] = approved.origin_subject.name
            ws[f'{ORIGIN_SUBJECTS_TABLE_CREDITS_COL}{current_row}'] = approved.origin_subject.credits
            ws[f'{ORIGIN_SUBJECTS_TABLE_GRADE_COL}{current_row}'] = approved.grade

            ws[f'{TARGET_SUBJECTS_TABLE_CODE_COL}{current_row}'] = approved.target_subject.code
            ws[f'{TARGET_SUBJECTS_TABLE_NAME_COL}{current_row}'] = approved.target_subject.name
            ws[f'{TARGET_SUBJECTS_TABLE_CREDITS_COL}{current_row}'] = approved.target_subject.credits
            ws[f'{TARGET_SUBJECTS_TABLE_GRADE_COL}{current_row}'] = approved.grade
            current_row += 1
        
